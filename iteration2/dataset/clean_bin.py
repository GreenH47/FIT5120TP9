from openpyxl import load_workbook

# Load the Excel workbook
workbook = load_workbook('waste_bin.xlsx')

# Select the appropriate sheet in the workbook (replace 'Sheet1' with the actual sheet name)
sheet = workbook['Sheet1']

# Iterate through each row in the sheet and delete the specified value
for row in sheet.iter_rows():
    for cell in row:
        if cell.value == "What couldn’t you put into the bin?":
            cell.value = ""

# Save the updated workbook
workbook.save('waste_bin_updated.xlsx')
